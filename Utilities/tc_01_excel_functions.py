""""
TestCase_excel_functions.py
Program-Python Program which will read and write your Excel file
"""

from openpyxl import load_workbook

class HRM_Excel_Functions:

    def __init__(self, file_name, sheet_number):
        self.file = file_name
        self.sheet = sheet_number

    #fetch the row count of the Excel file
    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row

    #fetch the column count of the Excel file
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column

    #read the data from the Excel file
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.cell(row=row_number, column=column_number).value

    #write the data to the Excel file
    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.file)
